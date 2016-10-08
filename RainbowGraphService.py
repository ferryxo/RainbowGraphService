from flask import Flask, request, render_template, jsonify
from flask_api import status
from flask_cors import CORS, cross_origin
import uuid
import json
import xlrd
import openpyxl
import numpy as np
import csv
import logging

formatter = logging.Formatter('%(asctime)s %(levelname)s %(message)s')

# first file logger
instructor_logger = logging.getLogger('instructor_logger')
hdlr_1 = logging.FileHandler('instructor.log')
hdlr_1.setFormatter(formatter)
instructor_logger.addHandler(hdlr_1)
instructor_logger.setLevel(logging.INFO)

# second file logger
debug_logger = logging.getLogger('debug_logger')
hdlr_2 = logging.FileHandler('debug.log')
hdlr_2.setFormatter(formatter)
debug_logger.addHandler(hdlr_2)
debug_logger.setLevel(logging.ERROR)

app = Flask(__name__)
CORS(app)

configTable = {}


@app.route('/file-upload', methods=['POST'])
@cross_origin()
def file_upload():
    file = request.files['file']

    #check type of the files, find a way to check the header
    if file.filename.endswith(".csv"):
        #log the headers
        header1 = file.readline()
        header2 = file.readline()
        instructor_logger.info(header1)
        instructor_logger.info(header2)

        if "Assignment =" in header1 and "Time =" in header2:
            #assume this is CPR file, to be confirmed


            best = 10
            worst = 0
            reader = csv.reader(file, dialect="excel")

            cpr_data = list(reader)

            authors_scores = {}
            prev_author = None
            prev_author_name = None
            current_author_scores = []
            list_of_author_json_conf = []

            for row in range(1, len(cpr_data)):
                author_id = cpr_data[row][0]
                author_name = cpr_data[row][2]
                reviewer = cpr_data[row][4]
                cpi = cpr_data[row][5]
                current_score = cpr_data[row][len(cpr_data[row])-2]
                primary_val = cpr_data[row][3].split("/")[0]

                if reviewer == '':
                    continue

                if author_id == prev_author or prev_author == None:
                    #add this reviewer and the score to this author
                    current_author_scores.append(float(current_score))
                elif prev_author != None:
                    #calculate the avg score & standard dev. for this author
                    authors_scores[prev_author] = current_author_scores

                    if "," in prev_author_name:
                        firstname = prev_author_name.split(",")[1] if len(prev_author_name.split(","))>1 else "-"
                        lastname = prev_author_name.split(",")[0]
                    else:
                        firstname = prev_author_name.split(" ")[1] if len(prev_author_name.split(" "))>1 else "-"
                        lastname = prev_author_name.split(" ")[0]

                    element_prev_author = {
                        "first_name": firstname,
                        "last_name": lastname,
                        "column_url": "",
                        "primary_value": float(primary_val),
                        "secondary_value": np.std(current_author_scores),
                        "values": current_author_scores
                    }
                    list_of_author_json_conf.append(element_prev_author)
                    current_author_scores = [float(current_score)]
                prev_author = author_id
                prev_author_name = author_name
        else:
            debug_logger.error("someone uploaded an unsupported csv file")
            return jsonify(error="Uploaded file is currently not supported"), status.HTTP_415_UNSUPPORTED_MEDIA_TYPE

    elif file.filename.endswith(".xls"):
        best = 7
        worst = 0
        book = xlrd.open_workbook(file_contents=file.read())
        data_sheet = book.sheet_by_index(0)

        #log the headers
        header1 = data_sheet.cell(0,0)  # 1st row
        instructor_logger.info(header1)
        header1 = data_sheet.cell(2,0)  # 2nd row
        instructor_logger.info(header1)
        header1 = data_sheet.cell(3,0)  # 3rd row
        instructor_logger.info(header1)

        if data_sheet.cell(4,0).value == "Paper Author" and data_sheet.cell(4,1).value == "Reviewer":
            #handle sword data
            authors_scores={}
            list_of_author_json_conf = []

            prev_author = None
            current_author_scores = []
            peer_given_holistic_score = {}
            for row in range(5, data_sheet.nrows):
                author = data_sheet.cell(row,0).value
                reviewer = data_sheet.cell(row,1).value

                if authors_scores.get(author) == None :
                    authors_scores[author] = {}

                #average scores all dimensions
                divider = 0
                score_sum = 0
                for col in range(2, data_sheet.ncols):
                    if not data_sheet.cell(row, col).value == '':
                        score_sum += float(data_sheet.cell(row, col).value)
                        divider += 1

                authors_scores[author][reviewer] = score_sum / float(divider)

                if prev_author == author or prev_author == None:
                    current_author_scores.append(authors_scores[author][reviewer])
                elif prev_author != None:
                    element_prev_author = {
                        "first_name": prev_author,
                        "last_name": "",
                        "column_url": "",
                        "primary_value": np.mean(current_author_scores),
                        "secondary_value": np.std(current_author_scores),
                        "values": current_author_scores
                    }
                    list_of_author_json_conf.append(element_prev_author)
                    current_author_scores = [authors_scores[author][reviewer]]
                prev_author = author
        else:
            debug_logger.error("someone uploaded an unsupported xls file")
            return jsonify(error="Uploaded file is currently not supported"), status.HTTP_415_UNSUPPORTED_MEDIA_TYPE

    elif file.filename.endswith(".xlsx"):
        debug_logger.error("someone uploaded an xlsx file")
        return jsonify(error="Uploaded file is currently not supported"), status.HTTP_415_UNSUPPORTED_MEDIA_TYPE
    else:
        debug_logger.error("someone uploaded unknown file (" + file.filename + ")" )
        return jsonify(error="Uploaded file is currently not supported"), status.HTTP_415_UNSUPPORTED_MEDIA_TYPE

    data = sorted(list_of_author_json_conf, key=lambda k: k['primary_value'], reverse=True)

    config = {
        "metadata": {
                    "primary-value-label": "rate average",
                    "higher_primary_value_better": True,
                    "values-label": "ranks",
                    "higher_values_better": False,
                    "best-value-possible": best,
                    "worst-value-possible": worst,
                    "y-axis-label": "Rate Average",
                    "x-axis-label": "Students",
                    "color-scheme": "5b",
                    "secondary-value-label": "variance"
            },
            "data": data
        }

    return jsonify([config]), status.HTTP_200_OK

@app.route('/instructor', methods=['GET'])
@cross_origin()
def instructor():
    #check type of the files
    return render_template('instructor.html')

@app.route('/configure', methods=['GET', 'POST'])
@cross_origin()
def index():
    if request.method == 'GET':
        return render_template('input_form.html')

    try:
        with open('configTable.json', 'r') as f:
            configTable = json.load(f)
    # if the file is empty the ValueError will be thrown
    except:
        configTable = {}

    id = uuid.uuid4();
    configTable[id.urn[9:]] = request.json

    with open('configTable.json', 'w+') as f:
        json.dump(configTable, f)

    return jsonify(url="http://peerlogic.csc.ncsu.edu/rainbowgraph/viz/" + id.urn[9:])

@app.route('/viz/<id>', methods=['GET', 'DELETE'])
@cross_origin()
def visualize(id):

 # load from file:
 try:
   with open('configTable.json', 'r+') as f:
       configTable = json.load(f)
    # if the file is empty the ValueError will be thrown
 except ValueError:
   configTable = {}
 config = configTable.get(id)

 if request.method == 'DELETE':
    f.close()
    configTable.pop(id, None)
    with open('configTable.json', 'w+') as f:
        json.dump(configTable, f)
    return "", status.HTTP_200_OK
 else:
     if config == None:
         return jsonify(error="Shoot.. I couldn't find the config data")
     else:
        return render_template('index.html', json_data = json.dumps(config))



if __name__ == '__main__':
    app.run(host='0.0.0.0', port=3005, threaded=True)
